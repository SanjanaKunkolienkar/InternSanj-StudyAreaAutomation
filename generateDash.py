# get output of all counties (red+blue) and (red+blue+green); only buses selected in counties - input to this function

from config.definitions import ROOT_DIR
import warnings
import os
import pandas as pd
import extractbuses as eb
import getcounty as gc
import psse_get_data as getdata
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image

def generate_dashboard(county_flip, county, buses_nl_v_dfax, filename, SA_county, POI_bus):
    county_blue = county_flip
    county_red = county

    county = set(county_red) | set(county_blue)
    county_df = pd.DataFrame(county, columns=['List of Study Area Counties'])


    buses_red = eb.extract_from_counties(county_red)
    buses_blue = eb.extract_from_counties(county_blue)

    buses_red_blue = buses_red + buses_blue
    #Calculating data for buses that are sensitive
    bus_df_sensitive = getdata.main(filename, buses_nl_v_dfax)
    total_gen_s = bus_df_sensitive['Gen MW'].sum()
    total_load_s = bus_df_sensitive['Load MW'].sum()
    tran_buses_s = bus_df_sensitive['Type'].value_counts()[float(1)]
    gen_buses_s =  bus_df_sensitive['Type'].value_counts()[float(2)]

    dash_data_s = [['Total Generation "Capacity" (units)', total_gen_s], ['Total Load (units)', total_load_s],
                   ['Total Number of Transmission Buses', tran_buses_s],
                   ['Total Number of Generation Buses', gen_buses_s],
                   ['Study County Name', SA_county],
                   ['POI', POI_bus]]
    dash_df_s = pd.DataFrame(dash_data_s, columns=['Summary', ''])


    transbus_df_sensitive = bus_df_sensitive[bus_df_sensitive['Type'] == 1]
    transbus_df_sensitive = transbus_df_sensitive.drop('Gen MW', axis=1)
    genbus_df_sensitive = bus_df_sensitive[bus_df_sensitive['Type'] == 2] #(bus_df_sensitive['Gen MW'].notna()) & (bus_df_sensitive['Gen MW'].apply(has_six_characters))
    genbus_df_sensitive = genbus_df_sensitive.drop('Load MW', axis=1)


    # Calculating data for all buses
    bus_df_red_blue = getdata.main(filename, buses_red_blue)
    total_gen_a = bus_df_red_blue['Gen MW'].sum()
    total_load_a = bus_df_red_blue['Load MW'].sum()
    tran_buses_a = bus_df_red_blue['Type'].value_counts()[float(1)]
    gen_buses_a = bus_df_red_blue['Type'].value_counts()[float(2)]

    dash_data_a = [['Total Generation "Capacity" (units)', total_gen_a], ['Total Load (units)', total_load_a],
                   ['Total Number of Transmission Buses', tran_buses_a],
                   ['Total Number of Generation Buses', gen_buses_a],
                   ['Study County Name', SA_county],
                   ['POI', POI_bus]]
    dash_df_a = pd.DataFrame(dash_data_a, columns=['Summary', ''])

    transbus_df_red_blue = bus_df_red_blue[bus_df_red_blue['Type'] == 1]
    transbus_df_red_blue = transbus_df_red_blue.drop('Gen MW', axis=1)
    genbus_df_red_blue = bus_df_red_blue[bus_df_red_blue['Type'] == 2] #(bus_df_red_blue['Gen MW'].notna()) & (bus_df_red_blue['Gen MW'].apply(has_six_characters))
    genbus_df_red_blue = genbus_df_red_blue.drop('Load MW', axis=1)



    filepath = os.path.join(ROOT_DIR, 'Input Data\SSWGCase\\', filename, 'Dashboard.xlsx')
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        #sensitive buses
        transbus_df_sensitive.to_excel(writer, startcol=0, startrow=3, sheet_name='Sheet1', index=False)
        genbus_df_sensitive.to_excel(writer, startcol=5, startrow=3, sheet_name='Sheet1', index=False)
        dash_df_s.to_excel(writer, startcol=10, startrow=3, sheet_name='Sheet1', index=False)
        #all buses
        transbus_df_red_blue.to_excel(writer, startcol=0, startrow=3, sheet_name='Sheet2', index=False)
        genbus_df_red_blue.to_excel(writer, startcol=5, startrow=3, sheet_name='Sheet2', index=False)
        dash_df_a.to_excel(writer, startcol=10, startrow=3, sheet_name='Sheet2', index=False)


    #Change sheet names to indicate sensitive buses and all buses in red + blue
    workbook = openpyxl.load_workbook(filepath)
    all_sheets = writer.sheets# Adding formats for header row.

    #renaming sheets
    for sheet in all_sheets:
        # Access the sheet you want to rename
        sheet_to_rename = workbook[sheet]
        print(sheet_to_rename)

        # Determine the new sheet name
        if sheet_to_rename.title == 'Sheet1':
            new_sheet_name = 'Sensitive buses'  # You can modify this pattern as needed
            header = 'Sensitive bus data in the output county map'
            image_name = 'CountyMap_Sensitive.jpg'
            # Update values in column2 based on values in column1
            county_df.loc[county_df['List of Study Area Counties'].isin(county_red), 'Legend'] = 'FFFF0000'
            county_df.loc[county_df['List of Study Area Counties'].isin(county_blue), 'Legend'] = 'FF00FFFF'
            county_df.loc[county_df['List of Study Area Counties'] == SA_county.lower(), 'Legend'] = 'FF00008B'
        if sheet_to_rename.title == 'Sheet2':
            new_sheet_name = 'All buses'  # You can modify this pattern as needed
            header = 'All bus data in the output county map'
            image_name = 'CountyMap_All.jpg'
            # Update values in column2 based on values in column1
            county_df.loc[county_df['List of Study Area Counties'].isin(county_red), 'Legend'] = 'FFFF0000'
            county_df.loc[county_df['List of Study Area Counties'].isin(county_blue), 'Legend'] = 'FFFF0000'
            county_df.loc[county_df['List of Study Area Counties'] == SA_county.lower(), 'Legend'] = 'FF00008B'

        # Rename the sheet
        sheet_to_rename.title = new_sheet_name

        # Define the range of cells to merge (e.g., A1 to B2)

        start_cell_main = 'A1'
        end_cell_main = 'L1'

        start_cell_t = 'A3'
        end_cell_t = 'D3'

        start_cell_g = 'F3'
        end_cell_g = 'I3'

        start_cell_d = 'K4'
        end_cell_d = 'L4'

        sheet = workbook[new_sheet_name]
        sheet[start_cell_t].value = 'Transmission Buses'
        sheet[start_cell_g].value = 'Generator Buses'
        sheet[start_cell_main].value = header

        sheet[start_cell_t].font = Font(bold=True)
        sheet[start_cell_g].font = Font(bold=True)
        sheet[start_cell_main].font = Font(bold=True)
        sheet['K12'].value = 'List of Study Area Counties'
        sheet['K12'].font = Font(bold=True)

        sheet[start_cell_t].alignment = Alignment(horizontal='center', vertical='center')
        sheet[start_cell_g].alignment = Alignment(horizontal='center', vertical='center')
        sheet[start_cell_main].alignment = Alignment(horizontal='center', vertical='center')

        # Merge the specified range of cells
        sheet.merge_cells(f'{start_cell_t}:{end_cell_t}')
        sheet.merge_cells(f'{start_cell_g}:{end_cell_g}')
        sheet.merge_cells(f'{start_cell_d}:{end_cell_d}')
        sheet.merge_cells(f'{start_cell_main}:{end_cell_main}')

        row_idx = 13
        for _, row in county_df.iterrows():
            name = row['List of Study Area Counties']
            color_code = row['Legend']

            cell = sheet.cell(row=row_idx, column=11, value=name)  # +2 because Excel rows start from 1
            cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
            if cell.value == SA_county.lower():
                cell.font = Font(color='FFFFFFFF')
            row_idx = row_idx + 1

        image_path = os.path.join(ROOT_DIR, 'Input Data\SSWGCase\\', filename, image_name)
        img = Image(image_path)  # Replace with the actual image file path
        sheet.add_image(img, 'N4')  # Add the image to cell A1 or the desired cell
        # Adjust the size of the image
        img.width = 800  # Specify the width in pixels
        img.height = 600  # Specify the height in pixels


    # Save the changes to the workbook
    workbook.save(filepath)

def has_six_characters(value):
    return len(str(value)) == 6
def main():
    generate_dashboard(county_flip, county, buses_nl_v_dfax, filename, SA_county, POI_bus)

if __name__ == "__main__":
    main()